from openpyxl import load_workbook
from tools.read_config import ReadConfig
from tools.project_path import *
from tools.get_data import GetData


class DoExcel(object):
    @staticmethod
    def get_data(file_name):
        wb = load_workbook(file_name)
        mode = eval(ReadConfig.get_config(test_config_path, 'MODE', 'mode'))
        test_data = []
        for key in mode:
            sheet = wb[key]
            if mode[key] == 'all':
                for i in range(2, sheet.max_row + 1):
                    row_data = dict()
                    row_data['case_id'] = sheet.cell(i, 1).value
                    row_data['url'] = sheet.cell(i, 2).value
                    # if sheet.cell(i, 3).value.find('${tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${tel}', str(tel))
                    #     tel = tel + 1
                    # elif sheet.cell(i, 3).value.find('${admin_tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${admin_tel}', str(getattr(GetData, 'admin_tel')))
                    # elif sheet.cell(i, 3).value.find('${loan_member_id}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${loan_member_id}', str(getattr(GetData, 'loan_member_id')))
                    # elif sheet.cell(i, 3).value.find('${normal_tel}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${normal_tel}', str(getattr(GetData, 'normal_tel')))
                    # elif sheet.cell(i, 3).value.find('${memberID}') != -1:
                    #     row_data['data'] = sheet.cell(i, 3).value.replace('${memberID}', str(getattr(GetData, 'memberID')))
                    # else:
                    #     row_data['data'] = sheet.cell(i, 3).value
                    row_data['data'] = sheet.cell(i, 3).value
                    row_data['title'] = sheet.cell(i, 4).value
                    row_data['http_method'] = sheet.cell(i, 5).value
                    row_data['header'] = sheet.cell(i, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
            else:
                for case_id in mode[key]:
                    row_data = dict()
                    row_data['case_id'] = sheet.cell(case_id+1, 1).value
                    row_data['url'] = sheet.cell(case_id+1, 2).value
                    row_data['data'] = sheet.cell(case_id+1, 3).value
                    row_data['title'] = sheet.cell(case_id+1, 4).value
                    row_data['http_method'] = sheet.cell(case_id+1, 5).value
                    row_data['header'] = sheet.cell(case_id+1, 6).value
                    row_data['sheet_name'] = key
                    test_data.append(row_data)
        return test_data

    @staticmethod
    def write_back(file_name, sheet_name, i, result, TestResult):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = result
        sheet.cell(i, 8).value = TestResult
        wb.save(file_name)


if __name__ == '__main__':
    res = DoExcel.get_data(test_data_path)
    print(len(res))

